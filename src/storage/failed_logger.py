from pathlib import Path
from datetime import datetime
from typing import TYPE_CHECKING

from openpyxl import Workbook, load_workbook

if TYPE_CHECKING:
    from ..tools import ColorfulConsole

__all__ = ["FailedLogger"]


class FailedLogger:
    """失败链接记录器"""
    
    def __init__(self, root: Path, console: "ColorfulConsole"):
        self.console = console
        self.path = root.joinpath("failed_links.xlsx")
        self.book = None
        self.sheet = None
        
    async def __aenter__(self):
        self.book = load_workbook(self.path) if self.path.exists() else Workbook()
        self.sheet = self.book.active
        self._init_title()
        return self
        
    async def __aexit__(self, exc_type, exc_val, exc_tb):
        if self.book:
            self.book.save(self.path)
            self.book.close()
            
    def _init_title(self):
        if not self.sheet["A1"].value:
            headers = ["时间", "链接", "失败原因", "类型"]
            for col, value in enumerate(headers, start=1):
                self.sheet.cell(row=1, column=col, value=value)
                
    async def log_failed_link(self, url: str, reason: str, link_type: str = "账号"):
        """记录失败的链接"""
        try:
            data = [
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                url,
                reason,
                link_type
            ]
            self.sheet.append(data)
        except Exception as e:
            self.console.warning(f"记录失败链接时出错: {e}")